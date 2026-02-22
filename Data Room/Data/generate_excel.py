#!/usr/bin/env python3
"""Generate Home Depot Equipment Rental Database Excel file"""

import json
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system("pip3 install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Load data
script_dir = os.path.dirname(os.path.abspath(__file__))
with open(os.path.join(script_dir, 'raw_data.json'), 'r') as f:
    raw_data = json.load(f)

with open(os.path.join(script_dir, 'equipment_specs.json'), 'r') as f:
    equipment_specs = json.load(f)

# Create workbook
wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="F96302", end_color="F96302", fill_type="solid")  # Home Depot orange
alt_fill = PatternFill(start_color="FFF3E6", end_color="FFF3E6", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ==================== SHEET 1: Summary ====================
ws_summary = wb.active
ws_summary.title = "Summary"

summary_data = [
    ["Home Depot Equipment Rental Database"],
    [""],
    ["Collection Date", "February 2026"],
    ["Total Unique Stores", "263"],
    ["States Covered", "TX, FL, NY"],
    ["Cities Searched", "70"],
    [""],
    ["Store Breakdown"],
    ["Texas", "~120 stores (25 cities)"],
    ["Florida", "~90 stores (25 cities)"],
    ["New York", "~53 stores (20 cities)"],
    [""],
    ["Equipment Categories"],
    ["Earth-Moving Equipment", "Mini Excavators, Skid Steers, Mini Skid Steers, TLB"],
    ["Aerial & Lifting Equipment", "Boom Lifts, Scissor Lifts, Forklifts, Telehandlers"],
    [""],
    ["Key Finding"],
    ["All large equipment shares the same 263-store network"],
    ["Once a store carries Mini Excavators, it carries all large equipment types"],
]

for row_idx, row_data in enumerate(summary_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = Font(bold=True, size=16)
        elif value in ["Store Breakdown", "Equipment Categories", "Key Finding"]:
            cell.font = Font(bold=True, size=12)

ws_summary.column_dimensions['A'].width = 35
ws_summary.column_dimensions['B'].width = 60

# ==================== SHEET 2: Stores ====================
ws_stores = wb.create_sheet("Stores")

# Headers
store_headers = ["Branch Name", "Address", "City", "State", "Phone", "Store Number"]
for col_idx, header in enumerate(store_headers, 1):
    cell = ws_stores.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Data
branches = raw_data.get('branches', {})
row_idx = 2
for branch_name, details in sorted(branches.items()):
    # Extract store number from branch name
    store_num = ""
    if "#" in branch_name:
        store_num = branch_name.split("#")[-1].strip()

    row_data = [
        branch_name,
        details.get('address', ''),
        details.get('city', ''),
        details.get('state', ''),
        details.get('phone', ''),
        store_num
    ]

    for col_idx, value in enumerate(row_data, 1):
        cell = ws_stores.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = alt_fill

    row_idx += 1

# Adjust column widths
ws_stores.column_dimensions['A'].width = 35
ws_stores.column_dimensions['B'].width = 45
ws_stores.column_dimensions['C'].width = 20
ws_stores.column_dimensions['D'].width = 8
ws_stores.column_dimensions['E'].width = 18
ws_stores.column_dimensions['F'].width = 12

# ==================== SHEET 3: Equipment Models ====================
ws_equipment = wb.create_sheet("Equipment Models")

# Headers
equip_headers = ["Category", "Equipment Type", "Model", "Availability", "4-Hour Rate", "Daily Rate", "Weekly Rate", "4-Week Rate", "Key Specs"]
for col_idx, header in enumerate(equip_headers, 1):
    cell = ws_equipment.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

row_idx = 2
categories = equipment_specs.get('categories', {})

for category_name, category_data in categories.items():
    for equip_type, type_data in category_data.items():
        models = type_data.get('models', [])
        if not models:
            # Equipment type with no specific models yet
            row_data = [
                category_name,
                equip_type,
                "Various models available",
                "In-store/Delivery",
                "", "", "", "",
                type_data.get('verification_note', 'Same 263 stores')
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_equipment.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
            row_idx += 1
        else:
            for model in models:
                specs = model.get('specs', {})
                pricing = model.get('rental_pricing', {})

                # Determine availability
                availability = specs.get('Availability', 'In-store pickup')

                # Build specs string
                spec_items = []
                for k, v in specs.items():
                    if k != 'Availability':
                        spec_items.append(f"{k}: {v}")
                specs_str = "; ".join(spec_items[:3])  # First 3 specs

                row_data = [
                    category_name,
                    equip_type,
                    model.get('name', ''),
                    availability,
                    pricing.get('4_hours', ''),
                    pricing.get('1_day', ''),
                    pricing.get('1_week', ''),
                    pricing.get('4_weeks', ''),
                    specs_str
                ]

                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_equipment.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    if row_idx % 2 == 0:
                        cell.fill = alt_fill

                row_idx += 1

# Adjust column widths
ws_equipment.column_dimensions['A'].width = 25
ws_equipment.column_dimensions['B'].width = 20
ws_equipment.column_dimensions['C'].width = 35
ws_equipment.column_dimensions['D'].width = 18
ws_equipment.column_dimensions['E'].width = 12
ws_equipment.column_dimensions['F'].width = 12
ws_equipment.column_dimensions['G'].width = 12
ws_equipment.column_dimensions['H'].width = 12
ws_equipment.column_dimensions['I'].width = 50

# ==================== SHEET 4: Stores by State ====================
ws_by_state = wb.create_sheet("Stores by State")

# Group stores by state
stores_by_state = {"TX": [], "FL": [], "NY": []}
for branch_name, details in branches.items():
    state = details.get('state', '')
    if state in stores_by_state:
        stores_by_state[state].append({
            'branch': branch_name,
            'city': details.get('city', ''),
            'address': details.get('address', ''),
            'phone': details.get('phone', '')
        })

# Headers
state_headers = ["State", "City", "Branch Name", "Address", "Phone"]
for col_idx, header in enumerate(state_headers, 1):
    cell = ws_by_state.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

row_idx = 2
for state in ["TX", "FL", "NY"]:
    # Sort by city
    stores = sorted(stores_by_state[state], key=lambda x: x['city'])
    for store in stores:
        row_data = [state, store['city'], store['branch'], store['address'], store['phone']]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_by_state.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill
        row_idx += 1

ws_by_state.column_dimensions['A'].width = 8
ws_by_state.column_dimensions['B'].width = 20
ws_by_state.column_dimensions['C'].width = 35
ws_by_state.column_dimensions['D'].width = 45
ws_by_state.column_dimensions['E'].width = 18

# Save workbook
output_path = os.path.join(script_dir, 'homedepot_equipment_database.xlsx')
wb.save(output_path)
print(f"Excel file created: {output_path}")
print(f"Sheets: Summary, Stores ({len(branches)} rows), Equipment Models, Stores by State")
