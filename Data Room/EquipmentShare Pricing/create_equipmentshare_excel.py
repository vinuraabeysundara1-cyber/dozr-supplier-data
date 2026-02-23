#!/usr/bin/env python3
"""
Create EquipmentShare Equipment Pricing Excel File
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# All equipment data with pricing
aerial_data = [
    {"category": "Aerial Work Platforms", "equipment_type": "Electric Scissor Lift", "name": "Electric Scissor Lift, 19' Narrow", "price_daily": 222, "price_weekly": 443, "price_monthly": 886},
    {"category": "Aerial Work Platforms", "equipment_type": "Vertical Mast Lift", "name": "Vertical Mast Lift Single Operator, 20'", "price_daily": 342, "price_weekly": 685, "price_monthly": 1370},
    {"category": "Aerial Work Platforms", "equipment_type": "Telescopic Boom Lift", "name": "Telescopic Boom Lift, 65' - 67' IC", "price_daily": 833, "price_weekly": 1875, "price_monthly": 3750},
    {"category": "Aerial Work Platforms", "equipment_type": "Articulating Boom Lift", "name": "Articulating Boom Lift, 125' IC", "price_daily": 2076, "price_weekly": 5189, "price_monthly": 11675},
    {"category": "Aerial Work Platforms", "equipment_type": "Articulating Boom Lift", "name": "Articulating Boom Lift, 135' IC", "price_daily": 2230, "price_weekly": 5017, "price_monthly": 12541},
    {"category": "Aerial Work Platforms", "equipment_type": "Articulating Boom Lift", "name": "Articulating Boom Lift, 30' - 35' IC", "price_daily": 545, "price_weekly": 1362, "price_monthly": 2724},
    {"category": "Aerial Work Platforms", "equipment_type": "Articulating Boom Lift", "name": "Articulating Boom Lift, 40' - 45' IC", "price_daily": 703, "price_weekly": 1581, "price_monthly": 3162},
    {"category": "Aerial Work Platforms", "equipment_type": "Articulating Boom Lift", "name": "Articulating Boom Lift, 60' - 65' IC", "price_daily": 945, "price_weekly": 2127, "price_monthly": 4254},
    {"category": "Aerial Work Platforms", "equipment_type": "Articulating Boom Lift", "name": "Articulating Boom Lift, 80' - 85' IC", "price_daily": 1220, "price_weekly": 3051, "price_monthly": 6102},
    {"category": "Aerial Work Platforms", "equipment_type": "Atrium Lift", "name": "Atrium Lift, 60' - 65'", "price_daily": 2403, "price_weekly": 5406, "price_monthly": 12164},
    {"category": "Aerial Work Platforms", "equipment_type": "Atrium Lift", "name": "Atrium Lift, 70' - 80'", "price_daily": 2784, "price_weekly": 6960, "price_monthly": 15661},
    {"category": "Aerial Work Platforms", "equipment_type": "Electric Boom Lift", "name": "Electric Articulating Boom Lift, 30' - 34'", "price_daily": 649, "price_weekly": 1622, "price_monthly": 3244},
    {"category": "Aerial Work Platforms", "equipment_type": "Electric Boom Lift", "name": "Electric Articulating Boom Lift, 40' Narrow", "price_daily": 528, "price_weekly": 1319, "price_monthly": 2639},
    {"category": "Aerial Work Platforms", "equipment_type": "Electric Boom Lift", "name": "Electric Articulating Boom Lift, 45' Wide", "price_daily": 668, "price_weekly": 1671, "price_monthly": 3342},
    {"category": "Aerial Work Platforms", "equipment_type": "Electric Scissor Lift", "name": "Electric Scissor Lift, 10' - 12'", "price_daily": 127, "price_weekly": 254, "price_monthly": 508},
    {"category": "Aerial Work Platforms", "equipment_type": "Electric Scissor Lift", "name": "Electric Scissor Lift, 13' - 14'", "price_daily": 171, "price_weekly": 342, "price_monthly": 683},
    {"category": "Aerial Work Platforms", "equipment_type": "Electric Scissor Lift", "name": "Electric Scissor Lift, 19' Micro", "price_daily": 222, "price_weekly": 443, "price_monthly": 886},
    {"category": "Aerial Work Platforms", "equipment_type": "Electric Scissor Lift", "name": "Electric Scissor Lift, 19' Micro with Step-up Platform", "price_daily": 159, "price_weekly": 316, "price_monthly": 632},
    {"category": "Aerial Work Platforms", "equipment_type": "Electric Scissor Lift", "name": "Electric Scissor Lift, 20' Narrow", "price_daily": 216, "price_weekly": 431, "price_monthly": 862},
    {"category": "Aerial Work Platforms", "equipment_type": "Electric Scissor Lift", "name": "Electric Scissor Lift, 26' Micro", "price_daily": 272, "price_weekly": 544, "price_monthly": 1087}
]

earthmoving_data = [
    {"category": "Earthmoving", "equipment_type": "Mini Excavator", "name": "Mini Excavator 10,000 - 14,000 lbs", "price_daily": 708, "price_weekly": 1948, "price_monthly": 4383},
    {"category": "Earthmoving", "equipment_type": "Mini Excavator", "name": "Mini Excavator 7,000 - 9,000 lbs", "price_daily": 449, "price_weekly": 1236, "price_monthly": 2781},
    {"category": "Earthmoving", "equipment_type": "Compact Track Loader", "name": "Track Skid Loader 3,100 - 3,400 Lbs ROC", "price_daily": 653, "price_weekly": 1797, "price_monthly": 4043},
    {"category": "Earthmoving", "equipment_type": "Dumper", "name": "All-Terrain Dumper", "price_daily": 559, "price_weekly": 1537, "price_monthly": 3458},
    {"category": "Earthmoving", "equipment_type": "Backhoe", "name": "Backhoe Loader 110 - 115 Hp, Extendable Stick", "price_daily": 508, "price_weekly": 1649, "price_monthly": 4536},
    {"category": "Earthmoving", "equipment_type": "Backhoe", "name": "Backhoe Loader 110 - 115 Hp, Standard Stick", "price_daily": 458, "price_weekly": 1489, "price_monthly": 4094},
    {"category": "Earthmoving", "equipment_type": "Backhoe", "name": "Backhoe Loader 68 - 74 Hp, Extendable Stick", "price_daily": 653, "price_weekly": 1796, "price_monthly": 3591},
    {"category": "Earthmoving", "equipment_type": "Backhoe", "name": "Backhoe Loader 68 - 74 Hp, Standard Stick", "price_daily": 594, "price_weekly": 1634, "price_monthly": 3267},
    {"category": "Earthmoving", "equipment_type": "Backhoe", "name": "Backhoe Loader 90 - 99 Hp, Extendable Stick", "price_daily": 585, "price_weekly": 1755, "price_monthly": 3509},
    {"category": "Earthmoving", "equipment_type": "Backhoe", "name": "Backhoe Loader 90 - 99 Hp, Standard Stick", "price_daily": 594, "price_weekly": 1485, "price_monthly": 3713},
    {"category": "Earthmoving", "equipment_type": "Skid Steer", "name": "Electric Mini Skid Steer 700 - 900 lbs ROC", "price_daily": 842, "price_weekly": 2316, "price_monthly": 5210},
    {"category": "Earthmoving", "equipment_type": "Mini Excavator", "name": "Mini Excavator 15,000 - 20,000 lbs", "price_daily": 844, "price_weekly": 2110, "price_monthly": 4748},
    {"category": "Earthmoving", "equipment_type": "Mini Excavator", "name": "Mini Excavator 2,500 - 4,000 lbs", "price_daily": 438, "price_weekly": 1204, "price_monthly": 2709},
    {"category": "Earthmoving", "equipment_type": "Mini Excavator", "name": "Mini Excavator 4,000 - 4,900 lbs, Electric", "price_daily": 527, "price_weekly": 1588, "price_monthly": 3574},
    {"category": "Earthmoving", "equipment_type": "Mini Excavator", "name": "Mini Excavator 5,000 - 6,500 lbs", "price_daily": 513, "price_weekly": 1410, "price_monthly": 3172},
    {"category": "Earthmoving", "equipment_type": "Skid Steer", "name": "Ride-On Mini Skid Steer 800 - 1,100 Lbs ROC", "price_daily": 392, "price_weekly": 1176, "price_monthly": 2647},
    {"category": "Earthmoving", "equipment_type": "Trencher", "name": "Ride-On Trencher, 45 - 50 Hp", "price_daily": 648, "price_weekly": 1620, "price_monthly": 3646},
    {"category": "Earthmoving", "equipment_type": "Backhoe", "name": "Skip Loader 60-85 HP", "price_daily": 421, "price_weekly": 1264, "price_monthly": 2528},
    {"category": "Earthmoving", "equipment_type": "Dozer", "name": "Track Dozer 16,000 - 18,000 lbs, 65 - 80 Hp", "price_daily": 831, "price_weekly": 2287, "price_monthly": 5716},
    {"category": "Earthmoving", "equipment_type": "Dozer", "name": "Track Dozer 20,000 lbs, 85 - 95 Hp", "price_daily": 881, "price_weekly": 2423, "price_monthly": 6056}
]

material_handling_data = [
    {"category": "Forklifts & Material Handling", "equipment_type": "Telehandler", "name": "Telehandler 5,000 - 5,500 lbs, 18' - 20' Reach", "price_daily": 729, "price_weekly": 1640, "price_monthly": 3281},
    {"category": "Forklifts & Material Handling", "equipment_type": "Telehandler", "name": "Telehandler 10,000 lbs, 54' - 56' Reach", "price_daily": 1157, "price_weekly": 2892, "price_monthly": 5783},
    {"category": "Forklifts & Material Handling", "equipment_type": "Beam Clamp", "name": "Beam Clamp 2 Ton", "price_daily": 7, "price_weekly": 17, "price_monthly": 43},
    {"category": "Forklifts & Material Handling", "equipment_type": "Beam Clamp", "name": "Beam Clamp 3 Ton", "price_daily": 9, "price_weekly": 24, "price_monthly": 61},
    {"category": "Forklifts & Material Handling", "equipment_type": "Beam Clamp", "name": "Beam Clamp 5 Ton", "price_daily": 13, "price_weekly": 32, "price_monthly": 80},
    {"category": "Forklifts & Material Handling", "equipment_type": "Beam Trolley", "name": "Beam Trolley 2 Ton", "price_daily": 15, "price_weekly": 41, "price_monthly": 108},
    {"category": "Forklifts & Material Handling", "equipment_type": "Beam Trolley", "name": "Beam Trolley 3 Ton", "price_daily": 23, "price_weekly": 62, "price_monthly": 154},
    {"category": "Forklifts & Material Handling", "equipment_type": "Carry Deck Crane", "name": "Carry Deck Crane 18,000lbs Diesel", "price_daily": 1096, "price_weekly": 2467, "price_monthly": 5551},
    {"category": "Forklifts & Material Handling", "equipment_type": "Carry Deck Crane", "name": "Carry Deck Crane 30,000lbs Dual Fuel", "price_daily": 1655, "price_weekly": 3723, "price_monthly": 8376},
    {"category": "Forklifts & Material Handling", "equipment_type": "Carry Deck Crane", "name": "Carry Deck Crane 36,000lbs Diesel", "price_daily": 1519, "price_weekly": 3797, "price_monthly": 9494},
    {"category": "Forklifts & Material Handling", "equipment_type": "Carry Deck Crane", "name": "Carry Deck Crane 8,000lbs Dual Fuel", "price_daily": 1083, "price_weekly": 2706, "price_monthly": 6089},
    {"category": "Forklifts & Material Handling", "equipment_type": "Carry Deck Crane", "name": "Carry Deck Crane 9,000lbs Diesel", "price_daily": 728, "price_weekly": 1819, "price_monthly": 4547},
    {"category": "Forklifts & Material Handling", "equipment_type": "Compact Tracked Crane", "name": "Compact Tracked Crane 5,000-6,000 lbs.", "price_daily": 1339, "price_weekly": 4017, "price_monthly": 12051},
    {"category": "Forklifts & Material Handling", "equipment_type": "Compact Tracked Crane", "name": "Compact Tracked Crane 5,500 - 6,500 lbs.", "price_daily": 941, "price_weekly": 2824, "price_monthly": 8473},
    {"category": "Forklifts & Material Handling", "equipment_type": "Gantry Crane", "name": "Gantry Crane Aluminum 2 Ton", "price_daily": 298, "price_weekly": 819, "price_monthly": 1842},
    {"category": "Forklifts & Material Handling", "equipment_type": "Hand Truck", "name": "Hand Truck Powered Cordless", "price_daily": 28, "price_weekly": 56, "price_monthly": 112},
    {"category": "Forklifts & Material Handling", "equipment_type": "Telehandler", "name": "High Reach Telehandler 10,000 lbs, 75' Reach", "price_daily": 2155, "price_weekly": 7004, "price_monthly": 17510}
]

# Create Excel file
with pd.ExcelWriter('equipmentshare_equipment_pricing.xlsx', engine='openpyxl') as writer:
    # Sheet 1: Aerial Work Platforms
    df_aerial = pd.DataFrame(aerial_data)
    df_aerial = df_aerial.sort_values(['equipment_type', 'name'])
    df_aerial.to_excel(writer, sheet_name='Aerial Work Platforms', index=False)

    # Sheet 2: Earthmoving
    df_earth = pd.DataFrame(earthmoving_data)
    df_earth = df_earth.sort_values(['equipment_type', 'name'])
    df_earth.to_excel(writer, sheet_name='Earthmoving', index=False)

    # Sheet 3: Forklifts & Material Handling
    df_material = pd.DataFrame(material_handling_data)
    df_material = df_material.sort_values(['equipment_type', 'name'])
    df_material.to_excel(writer, sheet_name='Material Handling', index=False)

    # Sheet 4: All Equipment
    all_data = aerial_data + earthmoving_data + material_handling_data
    df_all = pd.DataFrame(all_data)
    df_all = df_all.sort_values(['category', 'equipment_type', 'name'])
    df_all.to_excel(writer, sheet_name='All Equipment', index=False)

    # Sheet 5: Pricing Summary
    summary_data = []
    for category_name, category_data in [
        ('Aerial Work Platforms', aerial_data),
        ('Earthmoving', earthmoving_data),
        ('Forklifts & Material Handling', material_handling_data)
    ]:
        df_temp = pd.DataFrame(category_data)
        summary_data.append({
            'Category': category_name,
            'Equipment Count': len(category_data),
            'Min Daily Rate': df_temp['price_daily'].min(),
            'Max Daily Rate': df_temp['price_daily'].max(),
            'Avg Daily Rate': round(df_temp['price_daily'].mean(), 2),
            'Min Weekly Rate': df_temp['price_weekly'].min(),
            'Max Weekly Rate': df_temp['price_weekly'].max(),
            'Avg Weekly Rate': round(df_temp['price_weekly'].mean(), 2),
            'Min Monthly Rate': df_temp['price_monthly'].min(),
            'Max Monthly Rate': df_temp['price_monthly'].max(),
            'Avg Monthly Rate': round(df_temp['price_monthly'].mean(), 2)
        })

    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(writer, sheet_name='Pricing Summary', index=False)

# Format the Excel file
wb = load_workbook('equipmentshare_equipment_pricing.xlsx')

# Format each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Header formatting
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Auto-fit columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Currency formatting for price columns
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            col_letter = get_column_letter(cell.column)
            header = ws[f'{col_letter}1'].value

            if header and 'price' in header.lower() or 'rate' in str(header).lower():
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = '$#,##0'

    # Enable auto-filter
    ws.auto_filter.ref = ws.dimensions

wb.save('equipmentshare_equipment_pricing.xlsx')
print("✅ Excel file created: equipmentshare_equipment_pricing.xlsx")
print(f"   - 20 Aerial Work Platforms")
print(f"   - 20 Earthmoving equipment")
print(f"   - 17 Forklifts & Material Handling")
print(f"   - Total: 57 equipment items with complete pricing")
