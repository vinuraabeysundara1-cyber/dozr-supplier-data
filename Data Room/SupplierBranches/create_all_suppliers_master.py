#!/usr/bin/env python3
"""
Create All Suppliers Master Database
Combines branch data from:
1. Herc Rentals (565 branches)
2. Sunbelt Rentals (707+ OMS branches)
3. Sunstate Equipment (216 web branches)
4. EquipmentShare (289 web branches)

Outputs: all_suppliers_master_database.xlsx
"""

import json
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = "/Users/vinuraabeysundara/Desktop/ICG/DOZR/SupplierBranches"

# Supplier colors
COLORS = {
    'Herc Rentals': 'FFC107',      # Amber
    'Sunbelt Rentals': 'FF5722',   # Deep Orange
    'Sunstate Equipment': '4CAF50', # Green
    'EquipmentShare': '2196F3'     # Blue
}

def extract_branch_number(branch_name):
    """Extract branch number from name like 'Herc Rentals (9422) - Kansas City MO'"""
    if not branch_name:
        return None
    match = re.search(r'\((\d+)\)', branch_name)
    if match:
        return match.group(1)
    # Also try standalone numbers at start
    match = re.search(r'^(\d+)\s*-', branch_name)
    if match:
        return match.group(1)
    return None

def load_herc_data():
    """Load Herc branches from master database"""
    branches = []
    try:
        wb = load_workbook(f"{BASE_DIR}/herc_master_database.xlsx", read_only=True)
        ws = wb['All Unique Branches']

        # Skip header
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue

            branch = {
                'supplier': 'Herc Rentals',
                'branch_number': str(row[0] or '').strip(),
                'branch_name': str(row[1] or '').strip(),
                'address': str(row[2] or '').strip(),
                'city': str(row[3] or '').strip(),
                'state': str(row[4] or '').strip(),
                'zip': str(row[5] or '').strip(),
                'phone': str(row[6] or '').strip(),
                'services': str(row[7] or '').strip(),
                'in_oms': row[9] == 'Yes' if len(row) > 9 else False,
                'oms_enabled': row[10] == 'Yes' if len(row) > 10 else False,
                'source': str(row[11] or '').strip() if len(row) > 11 else ''
            }
            branches.append(branch)

        wb.close()
        print(f"Loaded {len(branches)} Herc branches")
    except Exception as e:
        print(f"Error loading Herc: {e}")

    return branches

def load_sunbelt_data():
    """Load Sunbelt branches from OMS + web data"""
    branches = []
    seen = set()

    # Load OMS data first
    try:
        with open(f"{BASE_DIR}/sunbelt_oms_branches.json", 'r') as f:
            data = json.load(f)

        for b in data.get('branches', []):
            branch_name = b.get('branch_name', '')
            branch_num = extract_branch_number(branch_name)
            key = branch_num or branch_name

            if key in seen:
                continue
            seen.add(key)

            branch = {
                'supplier': 'Sunbelt Rentals',
                'branch_number': branch_num or '',
                'branch_name': branch_name,
                'address': b.get('address', ''),
                'city': b.get('city', ''),
                'state': b.get('state', ''),
                'zip': b.get('zip', ''),
                'phone': b.get('phone', ''),
                'services': b.get('services', ''),
                'in_oms': True,
                'oms_enabled': b.get('oms_enabled', False),
                'source': 'oms'
            }
            branches.append(branch)

        print(f"Loaded {len(branches)} Sunbelt OMS branches")
    except Exception as e:
        print(f"Error loading Sunbelt OMS: {e}")

    # Add web branches not in OMS
    try:
        with open(f"{BASE_DIR}/sunbelt_all_branches.json", 'r') as f:
            data = json.load(f)

        web_added = 0
        for b in data.get('branches', []):
            branch_name = b.get('branch_name', '')
            key = branch_name

            if key in seen:
                continue
            seen.add(key)

            branch = {
                'supplier': 'Sunbelt Rentals',
                'branch_number': '',
                'branch_name': branch_name,
                'address': b.get('address', ''),
                'city': b.get('city', ''),
                'state': b.get('state', ''),
                'zip': b.get('zip', ''),
                'phone': b.get('phone', ''),
                'services': b.get('services', ''),
                'in_oms': False,
                'oms_enabled': False,
                'source': 'web'
            }
            branches.append(branch)
            web_added += 1

        print(f"Added {web_added} Sunbelt web branches (not in OMS)")
    except Exception as e:
        print(f"Error loading Sunbelt web: {e}")

    return branches

def load_sunstate_data():
    """Load Sunstate branches from OMS + web data"""
    branches = []
    seen = set()

    # Load OMS data first
    try:
        with open(f"{BASE_DIR}/sunstate_oms_branches.json", 'r') as f:
            data = json.load(f)

        for b in data.get('branches', []):
            branch_name = b.get('branch_name', '')
            key = branch_name

            if key in seen:
                continue
            seen.add(key)

            branch = {
                'supplier': 'Sunstate Equipment',
                'branch_number': extract_branch_number(branch_name) or '',
                'branch_name': branch_name,
                'address': b.get('address', ''),
                'city': b.get('city', ''),
                'state': b.get('state', ''),
                'zip': b.get('zip', ''),
                'phone': b.get('phone', ''),
                'services': b.get('services', ''),
                'in_oms': True,
                'oms_enabled': b.get('oms_enabled', False),
                'source': 'oms'
            }
            branches.append(branch)

        print(f"Loaded {len(branches)} Sunstate OMS branches")
    except Exception as e:
        print(f"Error loading Sunstate OMS: {e}")

    # Add web branches
    try:
        with open(f"{BASE_DIR}/sunstate_web_branches.json", 'r') as f:
            data = json.load(f)

        web_added = 0
        for b in data.get('branches', []):
            branch_name = b.get('branch_name', '')
            # Normalize for comparison
            key = branch_name.lower().strip()

            # Check if similar exists
            found = False
            for s in seen:
                if isinstance(s, str) and s.lower().strip() == key:
                    found = True
                    break

            if found:
                continue
            seen.add(branch_name)

            branch = {
                'supplier': 'Sunstate Equipment',
                'branch_number': '',
                'branch_name': branch_name,
                'address': b.get('address', ''),
                'city': b.get('city', ''),
                'state': b.get('state', ''),
                'zip': b.get('zip', ''),
                'phone': b.get('phone', ''),
                'services': b.get('services', ''),
                'in_oms': False,
                'oms_enabled': False,
                'source': 'web'
            }
            branches.append(branch)
            web_added += 1

        print(f"Added {web_added} Sunstate web branches (not in OMS)")
    except Exception as e:
        print(f"Error loading Sunstate web: {e}")

    return branches

def load_equipmentshare_data():
    """Load EquipmentShare branches from OMS + web data"""
    branches = []
    seen = set()

    # Load OMS data first
    try:
        with open(f"{BASE_DIR}/equipmentshare_oms_branches.json", 'r') as f:
            data = json.load(f)

        for b in data.get('branches', []):
            branch_name = b.get('branch_name', '')
            key = branch_name.lower().strip()

            if key in seen:
                continue
            seen.add(key)

            branch = {
                'supplier': 'EquipmentShare',
                'branch_number': extract_branch_number(branch_name) or '',
                'branch_name': branch_name,
                'address': b.get('address', ''),
                'city': b.get('city', ''),
                'state': b.get('state', ''),
                'zip': b.get('zip', ''),
                'phone': b.get('phone', ''),
                'services': b.get('services', ''),
                'in_oms': True,
                'oms_enabled': b.get('oms_enabled', False),
                'source': 'oms'
            }
            branches.append(branch)

        print(f"Loaded {len(branches)} EquipmentShare OMS branches")
    except Exception as e:
        print(f"Error loading EquipmentShare OMS: {e}")

    # Add web branches
    try:
        with open(f"{BASE_DIR}/equipmentshare_web_branches.json", 'r') as f:
            data = json.load(f)

        web_added = 0
        for b in data.get('branches', []):
            branch_name = b.get('branch_name', '')
            key = branch_name.lower().strip()

            if key in seen:
                continue
            seen.add(key)

            branch = {
                'supplier': 'EquipmentShare',
                'branch_number': '',
                'branch_name': branch_name,
                'address': b.get('address', ''),
                'city': b.get('city', ''),
                'state': b.get('state', ''),
                'zip': b.get('zip', ''),
                'phone': b.get('phone', ''),
                'services': b.get('services', ''),
                'in_oms': False,
                'oms_enabled': False,
                'source': 'web'
            }
            branches.append(branch)
            web_added += 1

        print(f"Added {web_added} EquipmentShare web branches (not in OMS)")
    except Exception as e:
        print(f"Error loading EquipmentShare web: {e}")

    return branches

def create_master_database(herc, sunbelt, sunstate, equipmentshare):
    """Create master database with all suppliers"""

    all_branches = herc + sunbelt + sunstate + equipmentshare

    # Create workbook
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: All Suppliers Combined
    ws1 = wb.active
    ws1.title = "All Branches"
    headers = ['Supplier', 'Branch #', 'Branch Name', 'Address', 'City', 'State', 'ZIP', 'Phone', 'Services', 'In OMS', 'OMS Enabled', 'Source']
    ws1.append(headers)

    header_fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    for col in range(1, len(headers) + 1):
        ws1.cell(row=1, column=col).font = header_font
        ws1.cell(row=1, column=col).fill = header_fill
        ws1.cell(row=1, column=col).border = thin_border

    for b in all_branches:
        row = [
            b.get('supplier', ''),
            b.get('branch_number', ''),
            b.get('branch_name', ''),
            b.get('address', ''),
            b.get('city', ''),
            b.get('state', ''),
            b.get('zip', ''),
            b.get('phone', ''),
            b.get('services', ''),
            'Yes' if b.get('in_oms') else 'No',
            'Yes' if b.get('oms_enabled') else 'No',
            b.get('source', '')
        ]
        ws1.append(row)

    # Set column widths
    ws1.column_dimensions['A'].width = 18
    ws1.column_dimensions['B'].width = 10
    ws1.column_dimensions['C'].width = 40
    ws1.column_dimensions['D'].width = 35
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 8
    ws1.column_dimensions['G'].width = 10
    ws1.column_dimensions['H'].width = 15
    ws1.column_dimensions['I'].width = 20

    # Sheet 2: Herc Rentals
    ws2 = wb.create_sheet("Herc Rentals")
    create_supplier_sheet(ws2, herc, 'Herc Rentals', COLORS['Herc Rentals'])

    # Sheet 3: Sunbelt Rentals
    ws3 = wb.create_sheet("Sunbelt Rentals")
    create_supplier_sheet(ws3, sunbelt, 'Sunbelt Rentals', COLORS['Sunbelt Rentals'])

    # Sheet 4: Sunstate Equipment
    ws4 = wb.create_sheet("Sunstate Equipment")
    create_supplier_sheet(ws4, sunstate, 'Sunstate Equipment', COLORS['Sunstate Equipment'])

    # Sheet 5: EquipmentShare
    ws5 = wb.create_sheet("EquipmentShare")
    create_supplier_sheet(ws5, equipmentshare, 'EquipmentShare', COLORS['EquipmentShare'])

    # Sheet 6: State Summary
    ws6 = wb.create_sheet("State Summary")
    create_state_summary(ws6, all_branches)

    # Sheet 7: Supplier Summary
    ws7 = wb.create_sheet("Summary")
    create_summary(ws7, herc, sunbelt, sunstate, equipmentshare)

    # Save
    output_path = f"{BASE_DIR}/all_suppliers_master_database.xlsx"
    wb.save(output_path)

    return {
        'total': len(all_branches),
        'herc': len(herc),
        'sunbelt': len(sunbelt),
        'sunstate': len(sunstate),
        'equipmentshare': len(equipmentshare),
        'in_oms': sum(1 for b in all_branches if b.get('in_oms')),
        'oms_enabled': sum(1 for b in all_branches if b.get('oms_enabled')),
        'output_path': output_path
    }

def create_supplier_sheet(ws, branches, supplier_name, color):
    """Create a sheet for a single supplier"""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    headers = ['Branch #', 'Branch Name', 'Address', 'City', 'State', 'ZIP', 'Phone', 'Services', 'In OMS', 'OMS Enabled']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).fill = header_fill

    for b in branches:
        row = [
            b.get('branch_number', ''),
            b.get('branch_name', ''),
            b.get('address', ''),
            b.get('city', ''),
            b.get('state', ''),
            b.get('zip', ''),
            b.get('phone', ''),
            b.get('services', ''),
            'Yes' if b.get('in_oms') else 'No',
            'Yes' if b.get('oms_enabled') else 'No'
        ]
        ws.append(row)

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20

def create_state_summary(ws, all_branches):
    """Create state summary sheet"""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")

    # Count by state and supplier
    state_counts = {}
    for b in all_branches:
        state = b.get('state', 'Unknown') or 'Unknown'
        supplier = b.get('supplier', 'Unknown')

        if state not in state_counts:
            state_counts[state] = {'Herc Rentals': 0, 'Sunbelt Rentals': 0, 'Sunstate Equipment': 0, 'EquipmentShare': 0, 'Total': 0}

        state_counts[state][supplier] += 1
        state_counts[state]['Total'] += 1

    headers = ['State', 'Herc', 'Sunbelt', 'Sunstate', 'EquipmentShare', 'Total']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).fill = header_fill

    for state in sorted(state_counts.keys()):
        counts = state_counts[state]
        ws.append([
            state,
            counts['Herc Rentals'],
            counts['Sunbelt Rentals'],
            counts['Sunstate Equipment'],
            counts['EquipmentShare'],
            counts['Total']
        ])

    # Add totals
    totals = ['TOTAL',
              sum(c['Herc Rentals'] for c in state_counts.values()),
              sum(c['Sunbelt Rentals'] for c in state_counts.values()),
              sum(c['Sunstate Equipment'] for c in state_counts.values()),
              sum(c['EquipmentShare'] for c in state_counts.values()),
              sum(c['Total'] for c in state_counts.values())]
    ws.append(totals)

    last_row = ws.max_row
    for col in range(1, len(headers) + 1):
        ws.cell(row=last_row, column=col).font = Font(bold=True)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

def create_summary(ws, herc, sunbelt, sunstate, equipmentshare):
    """Create summary sheet"""
    all_branches = herc + sunbelt + sunstate + equipmentshare

    summary_data = [
        ['ALL SUPPLIERS MASTER DATABASE', ''],
        ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M')],
        ['', ''],
        ['SUPPLIER TOTALS', ''],
        ['Herc Rentals', len(herc)],
        ['Sunbelt Rentals', len(sunbelt)],
        ['Sunstate Equipment', len(sunstate)],
        ['EquipmentShare', len(equipmentshare)],
        ['', ''],
        ['GRAND TOTAL', len(all_branches)],
        ['', ''],
        ['OMS STATUS', ''],
        ['In OMS Database', sum(1 for b in all_branches if b.get('in_oms'))],
        ['NOT in OMS (New)', sum(1 for b in all_branches if not b.get('in_oms'))],
        ['OMS Enabled', sum(1 for b in all_branches if b.get('oms_enabled'))],
        ['', ''],
        ['BREAKDOWN BY SUPPLIER', ''],
        ['', 'Total | In OMS | OMS Enabled | New'],
        ['Herc Rentals', f"{len(herc)} | {sum(1 for b in herc if b.get('in_oms'))} | {sum(1 for b in herc if b.get('oms_enabled'))} | {sum(1 for b in herc if not b.get('in_oms'))}"],
        ['Sunbelt Rentals', f"{len(sunbelt)} | {sum(1 for b in sunbelt if b.get('in_oms'))} | {sum(1 for b in sunbelt if b.get('oms_enabled'))} | {sum(1 for b in sunbelt if not b.get('in_oms'))}"],
        ['Sunstate Equipment', f"{len(sunstate)} | {sum(1 for b in sunstate if b.get('in_oms'))} | {sum(1 for b in sunstate if b.get('oms_enabled'))} | {sum(1 for b in sunstate if not b.get('in_oms'))}"],
        ['EquipmentShare', f"{len(equipmentshare)} | {sum(1 for b in equipmentshare if b.get('in_oms'))} | {sum(1 for b in equipmentshare if b.get('oms_enabled'))} | {sum(1 for b in equipmentshare if not b.get('in_oms'))}"],
    ]

    for row in summary_data:
        ws.append(row)

    # Style
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    for row_num in [4, 10, 12, 17]:
        ws.cell(row=row_num, column=1).font = Font(bold=True)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40

def main():
    print("=" * 70)
    print("CREATING ALL SUPPLIERS MASTER DATABASE")
    print("=" * 70)
    print()

    # Load all supplier data
    print("Loading supplier data...")
    herc = load_herc_data()
    sunbelt = load_sunbelt_data()
    sunstate = load_sunstate_data()
    equipmentshare = load_equipmentshare_data()

    print()

    # Create master database
    print("Creating master database...")
    results = create_master_database(herc, sunbelt, sunstate, equipmentshare)

    print()
    print("=" * 70)
    print("FINAL COUNTS")
    print("=" * 70)
    print(f"Herc Rentals:        {results['herc']:>6} branches")
    print(f"Sunbelt Rentals:     {results['sunbelt']:>6} branches")
    print(f"Sunstate Equipment:  {results['sunstate']:>6} branches")
    print(f"EquipmentShare:      {results['equipmentshare']:>6} branches")
    print("-" * 40)
    print(f"TOTAL:               {results['total']:>6} branches")
    print()
    print(f"In OMS Database:     {results['in_oms']:>6}")
    print(f"OMS Enabled:         {results['oms_enabled']:>6}")
    print()
    print(f"Output saved to: {results['output_path']}")
    print("=" * 70)

if __name__ == "__main__":
    main()
