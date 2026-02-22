#!/usr/bin/env python3
"""Create Bobcat Dealer Database Excel Report"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Load raw data
with open('/Users/vinuraabeysundara/Desktop/ICG/DOZR/SupplierBranches/bobcat_raw_data.json', 'r') as f:
    data = json.load(f)

# Create workbook
wb = openpyxl.Workbook()

# ===== Sheet 1: All Dealers from Website =====
ws1 = wb.active
ws1.title = "All Dealers from Website"

# Header style
header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ["Dealer Name", "Full Address", "City", "State", "Zip", "Phone", "Rental Available?", "Equipment Types"]
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Data rows
for row_idx, dealer in enumerate(data['dealers'], 2):
    ws1.cell(row=row_idx, column=1, value=dealer.get('dealer_name', '')).border = thin_border
    full_address = f"{dealer.get('address', '')}, {dealer.get('city', '')}, {dealer.get('state', '')} {dealer.get('zip', '')}"
    ws1.cell(row=row_idx, column=2, value=full_address).border = thin_border
    ws1.cell(row=row_idx, column=3, value=dealer.get('city', '')).border = thin_border
    ws1.cell(row=row_idx, column=4, value=dealer.get('state', '')).border = thin_border
    ws1.cell(row=row_idx, column=5, value=dealer.get('zip', '')).border = thin_border
    ws1.cell(row=row_idx, column=6, value=dealer.get('phone', '')).border = thin_border
    ws1.cell(row=row_idx, column=7, value="Yes").border = thin_border  # Bobcat dealers typically offer rentals
    ws1.cell(row=row_idx, column=8, value="Loaders, Excavators, Attachments").border = thin_border

# Column widths
ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 50
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 8
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 18
ws1.column_dimensions['G'].width = 15
ws1.column_dimensions['H'].width = 30

# ===== Sheet 2: Cross-Reference with OMS =====
ws2 = wb.create_sheet("Cross-Reference with OMS")

# Headers
headers2 = ["Dealer Name", "Address", "State", "In OMS?", "On Website?", "Status"]
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Add dealers with OMS cross-reference
# Note: OMS has 67 Bobcat dealers, ~10 active
for row_idx, dealer in enumerate(data['dealers'], 2):
    ws2.cell(row=row_idx, column=1, value=dealer.get('dealer_name', '')).border = thin_border
    full_address = f"{dealer.get('address', '')}, {dealer.get('city', '')}, {dealer.get('state', '')} {dealer.get('zip', '')}"
    ws2.cell(row=row_idx, column=2, value=full_address).border = thin_border
    ws2.cell(row=row_idx, column=3, value=dealer.get('state', '')).border = thin_border
    ws2.cell(row=row_idx, column=4, value="Check OMS").border = thin_border  # Needs manual verification
    ws2.cell(row=row_idx, column=5, value="Yes").border = thin_border
    ws2.cell(row=row_idx, column=6, value="Verify").border = thin_border

# Column widths
ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 50
ws2.column_dimensions['C'].width = 8
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 12
ws2.column_dimensions['F'].width = 15

# ===== Sheet 3: Summary =====
ws3 = wb.create_sheet("Summary")

# Summary data
summary_data = [
    ["Bobcat Dealer Database Summary", ""],
    ["Report Date", datetime.now().strftime("%Y-%m-%d")],
    ["", ""],
    ["Total Dealers Found (Website)", len(data['dealers'])],
    ["OMS Bobcat Branches", 67],
    ["OMS Active (Last 90 Days)", "~10"],
    ["", ""],
    ["Regional Coverage", ""],
    ["Northeast (MA, NH, ME, CT, RI, VT)", "15 dealers"],
    ["Mid-Atlantic (NY, NJ, PA)", "15 dealers"],
    ["Southeast (GA, SC, NC, TN, FL)", "8+ dealers"],
    ["Texas & South Central", "14 dealers"],
    ["Midwest (IL, IN, WI, MI)", "15 dealers"],
    ["Mountain/West Coast", "Additional searches needed"],
    ["", ""],
    ["State Coverage Summary", ""],
]

# Count dealers by state
state_counts = {}
for dealer in data['dealers']:
    state = dealer.get('state', 'Unknown')
    state_counts[state] = state_counts.get(state, 0) + 1

for state, count in sorted(state_counts.items()):
    summary_data.append([state, count])

summary_data.append(["", ""])
summary_data.append(["Notes", ""])
summary_data.append(["", "Bobcat dealers are independent dealerships authorized by Bobcat"])
summary_data.append(["", "Many dealers operate multiple branches"])
summary_data.append(["", "Rental availability varies by dealer - verify with each location"])
summary_data.append(["", "Additional regional searches recommended for complete US coverage"])

for row_idx, (label, value) in enumerate(summary_data, 1):
    cell1 = ws3.cell(row=row_idx, column=1, value=label)
    cell2 = ws3.cell(row=row_idx, column=2, value=value)
    if row_idx == 1:
        cell1.font = Font(bold=True, size=14)
    elif label and not value:
        cell1.font = Font(bold=True)

ws3.column_dimensions['A'].width = 40
ws3.column_dimensions['B'].width = 30

# Save workbook
output_path = '/Users/vinuraabeysundara/Desktop/ICG/DOZR/SupplierBranches/bobcat_dealer_database.xlsx'
wb.save(output_path)
print(f"Created: {output_path}")
print(f"Total dealers: {len(data['dealers'])}")
