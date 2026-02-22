#!/usr/bin/env python3
"""
Create comprehensive Excel report with all supplier branch data
Includes both OMS data and web extraction data
"""

import json
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Installing openpyxl...")
    os.system("pip3 install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = "/Users/vinuraabeysundara/Desktop/ICG/DOZR/SupplierBranches"

# Style definitions
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ENABLED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
DISABLED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WEB_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def load_json(filename):
    """Load JSON file from base directory"""
    filepath = os.path.join(BASE_DIR, filename)
    if os.path.exists(filepath):
        with open(filepath, 'r') as f:
            return json.load(f)
    return None

def style_header_row(ws, row_num, num_cols):
    """Apply header styling to a row"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

def add_oms_branch_sheet(wb, supplier_data, sheet_name):
    """Add a sheet with OMS branch data for a supplier"""
    ws = wb.create_sheet(title=sheet_name[:31])

    headers = [
        "Branch ID", "Branch Name", "Address", "City", "State",
        "Zip Code", "Phone", "Email", "OMS Enabled", "Hours (Mon-Fri)", "Source"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, len(headers))

    branches = supplier_data.get('branches', [])
    for row_idx, branch in enumerate(branches, 2):
        hours = branch.get('hours', {})
        mon_hours = hours.get('monday', 'N/A') if hours else 'N/A'

        ws.cell(row=row_idx, column=1, value=branch.get('branch_id', ''))
        ws.cell(row=row_idx, column=2, value=branch.get('branch_name', ''))
        ws.cell(row=row_idx, column=3, value=branch.get('address', ''))
        ws.cell(row=row_idx, column=4, value=branch.get('city', ''))
        ws.cell(row=row_idx, column=5, value=branch.get('state', ''))
        ws.cell(row=row_idx, column=6, value=branch.get('zip', ''))
        ws.cell(row=row_idx, column=7, value=branch.get('phone', ''))
        ws.cell(row=row_idx, column=8, value=branch.get('email', ''))

        oms_enabled = branch.get('oms_enabled', False)
        oms_cell = ws.cell(row=row_idx, column=9, value="Yes" if oms_enabled else "No")
        oms_cell.fill = ENABLED_FILL if oms_enabled else DISABLED_FILL

        ws.cell(row=row_idx, column=10, value=mon_hours if mon_hours else 'N/A')
        ws.cell(row=row_idx, column=11, value="OMS")

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = BORDER

    # Column widths
    column_widths = [20, 45, 35, 20, 8, 12, 15, 35, 12, 15, 8]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'
    return len(branches)

def add_web_branch_sheet(wb, web_data, sheet_name):
    """Add a sheet with web-extracted branch data"""
    ws = wb.create_sheet(title=sheet_name[:31])

    headers = [
        "Branch Name", "Address", "City", "State", "Zip Code",
        "Phone", "Services", "Source"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, len(headers))

    branches = web_data.get('branches', [])
    for row_idx, branch in enumerate(branches, 2):
        ws.cell(row=row_idx, column=1, value=branch.get('branch_name', ''))
        ws.cell(row=row_idx, column=2, value=branch.get('address', ''))
        ws.cell(row=row_idx, column=3, value=branch.get('city', ''))
        ws.cell(row=row_idx, column=4, value=branch.get('state', ''))
        ws.cell(row=row_idx, column=5, value=branch.get('zip', ''))

        # Clean phone number
        phone = branch.get('phone', '')
        if 'Questions?' in phone:
            phone = ''
        ws.cell(row=row_idx, column=6, value=phone)

        ws.cell(row=row_idx, column=7, value=branch.get('services', ''))

        source_cell = ws.cell(row=row_idx, column=8, value="Web")
        source_cell.fill = WEB_FILL

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = BORDER

    # Column widths
    column_widths = [45, 35, 20, 15, 12, 15, 25, 8]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'
    return len(branches)

def add_combined_sheet(wb, oms_data, web_data, sheet_name, supplier_name):
    """Add a sheet with combined OMS + Web data"""
    ws = wb.create_sheet(title=sheet_name[:31])

    headers = [
        "Branch Name", "Address", "City", "State", "Zip Code",
        "Phone", "Email", "OMS Enabled", "Services", "Source"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, len(headers))

    row_idx = 2

    # Add OMS branches first
    if oms_data:
        for branch in oms_data.get('branches', []):
            ws.cell(row=row_idx, column=1, value=branch.get('branch_name', ''))
            ws.cell(row=row_idx, column=2, value=branch.get('address', ''))
            ws.cell(row=row_idx, column=3, value=branch.get('city', ''))
            ws.cell(row=row_idx, column=4, value=branch.get('state', ''))
            ws.cell(row=row_idx, column=5, value=branch.get('zip', ''))
            ws.cell(row=row_idx, column=6, value=branch.get('phone', ''))
            ws.cell(row=row_idx, column=7, value=branch.get('email', ''))

            oms_enabled = branch.get('oms_enabled', False)
            oms_cell = ws.cell(row=row_idx, column=8, value="Yes" if oms_enabled else "No")
            oms_cell.fill = ENABLED_FILL if oms_enabled else DISABLED_FILL

            ws.cell(row=row_idx, column=9, value='')
            ws.cell(row=row_idx, column=10, value="OMS")

            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = BORDER
            row_idx += 1

    # Add Web branches (that might not be in OMS)
    oms_names = set()
    if oms_data:
        for b in oms_data.get('branches', []):
            name = b.get('branch_name', '').lower()
            city = b.get('city', '').lower()
            oms_names.add(f"{name}|{city}")

    if web_data:
        for branch in web_data.get('branches', []):
            # Check if this branch might already be in OMS
            name = branch.get('branch_name', '').lower()
            city = branch.get('city', '').lower()
            key = f"{name}|{city}"

            # Skip if likely duplicate
            is_duplicate = any(city in oms_key for oms_key in oms_names)

            if not is_duplicate:
                ws.cell(row=row_idx, column=1, value=branch.get('branch_name', ''))
                ws.cell(row=row_idx, column=2, value=branch.get('address', ''))
                ws.cell(row=row_idx, column=3, value=branch.get('city', ''))
                ws.cell(row=row_idx, column=4, value=branch.get('state', ''))
                ws.cell(row=row_idx, column=5, value=branch.get('zip', ''))

                phone = branch.get('phone', '')
                if 'Questions?' in phone:
                    phone = ''
                ws.cell(row=row_idx, column=6, value=phone)

                ws.cell(row=row_idx, column=7, value='')
                ws.cell(row=row_idx, column=8, value="N/A")
                ws.cell(row=row_idx, column=9, value=branch.get('services', ''))

                source_cell = ws.cell(row=row_idx, column=10, value="Web")
                source_cell.fill = WEB_FILL

                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).border = BORDER
                row_idx += 1

    # Column widths
    column_widths = [45, 35, 20, 8, 12, 15, 35, 12, 25, 8]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'
    return row_idx - 2

def add_summary_sheet(wb, stats):
    """Add summary sheet"""
    ws = wb.create_sheet(title="Summary", index=0)

    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value="Supplier Branch Data - Complete Report")
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')

    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=3, column=1, value="")

    # Headers
    headers = ["Supplier", "OMS Branches", "OMS Enabled", "Web Branches", "Combined Total", "Enable Rate", "Action"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    style_header_row(ws, 4, len(headers))

    # Data
    row = 5
    totals = {'oms': 0, 'enabled': 0, 'web': 0, 'combined': 0}

    for supplier, data in stats.items():
        ws.cell(row=row, column=1, value=supplier)
        ws.cell(row=row, column=2, value=data['oms'])

        enabled_cell = ws.cell(row=row, column=3, value=data['enabled'])
        enabled_cell.fill = ENABLED_FILL if data['enabled'] > 0 else DISABLED_FILL

        web_cell = ws.cell(row=row, column=4, value=data['web'])
        if data['web'] > 0:
            web_cell.fill = WEB_FILL

        ws.cell(row=row, column=5, value=data['combined'])

        rate = (data['enabled'] / data['oms'] * 100) if data['oms'] > 0 else 0
        ws.cell(row=row, column=6, value=f"{rate:.1f}%")

        action = "Complete" if rate == 100 else f"Enable {data['oms'] - data['enabled']} branches"
        action_cell = ws.cell(row=row, column=7, value=action)
        if rate < 100:
            action_cell.font = Font(color="FF0000", bold=True)

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = BORDER

        totals['oms'] += data['oms']
        totals['enabled'] += data['enabled']
        totals['web'] += data['web']
        totals['combined'] += data['combined']
        row += 1

    # Totals row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=totals['oms']).font = Font(bold=True)
    ws.cell(row=row, column=3, value=totals['enabled']).font = Font(bold=True)
    ws.cell(row=row, column=4, value=totals['web']).font = Font(bold=True)
    ws.cell(row=row, column=5, value=totals['combined']).font = Font(bold=True)

    overall_rate = (totals['enabled'] / totals['oms'] * 100) if totals['oms'] > 0 else 0
    ws.cell(row=row, column=6, value=f"{overall_rate:.1f}%").font = Font(bold=True)
    ws.cell(row=row, column=7, value=f"Enable {totals['oms'] - totals['enabled']} branches").font = Font(bold=True, color="FF0000")

    for col in range(1, len(headers) + 1):
        ws.cell(row=row, column=col).border = BORDER

    # Column widths
    column_widths = [20, 15, 15, 15, 15, 12, 25]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Legend
    row += 3
    ws.cell(row=row, column=1, value="Legend:").font = Font(bold=True)
    row += 1

    legend_cell = ws.cell(row=row, column=1, value="OMS Enabled")
    legend_cell.fill = ENABLED_FILL
    row += 1

    legend_cell = ws.cell(row=row, column=1, value="OMS Disabled")
    legend_cell.fill = DISABLED_FILL
    row += 1

    legend_cell = ws.cell(row=row, column=1, value="Web Extracted")
    legend_cell.fill = WEB_FILL

def main():
    print("Creating comprehensive Excel report...")
    print("=" * 50)

    # Load all data
    oms_data = {
        'Sunstate': load_json('sunstate_oms_branches.json'),
        'Herc': load_json('herc_oms_branches.json'),
        'Sunbelt': load_json('sunbelt_oms_branches.json'),
        'EquipmentShare': load_json('equipmentshare_oms_branches.json')
    }

    web_data = {
        'Sunstate': load_json('sunstate_web_branches.json'),
        'Sunbelt': load_json('sunbelt_all_branches.json'),
        'EquipmentShare': load_json('equipmentshare_web_branches.json')
    }

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Collect stats
    stats = {}

    for supplier in ['Sunstate', 'Herc', 'Sunbelt', 'EquipmentShare']:
        oms = oms_data.get(supplier)
        web = web_data.get(supplier)

        oms_count = len(oms.get('branches', [])) if oms else 0
        oms_enabled = oms.get('oms_enabled_count', 0) if oms else 0
        web_count = len(web.get('branches', [])) if web else 0

        stats[supplier] = {
            'oms': oms_count,
            'enabled': oms_enabled,
            'web': web_count,
            'combined': max(oms_count, web_count)  # Approximate unique
        }

        # Add OMS sheet
        if oms:
            count = add_oms_branch_sheet(wb, oms, f"{supplier} OMS")
            print(f"  {supplier} OMS: {count} branches")

        # Add Web sheet if available
        if web and web_count > 0:
            count = add_web_branch_sheet(wb, web, f"{supplier} Web")
            print(f"  {supplier} Web: {count} branches")

    # Add summary sheet
    add_summary_sheet(wb, stats)
    print("  Summary sheet created")

    # Save
    output_file = os.path.join(BASE_DIR, "All_Supplier_Branches_Complete_Report.xlsx")
    wb.save(output_file)

    print("\n" + "=" * 50)
    print(f"Report saved to: {output_file}")
    print("\nSheets included:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")

if __name__ == "__main__":
    main()
