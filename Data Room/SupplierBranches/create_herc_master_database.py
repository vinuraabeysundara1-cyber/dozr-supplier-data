#!/usr/bin/env python3
"""
Create Herc Rentals Master Database
Combines three data sources:
1. herc_complete_branch_database.xlsx (web scrape - 375 branches)
2. herc_gap_fill_branches.json (gap fill scrape - 80+ branches)
3. Herc-Rentals-19-Feb-26.csv (OMS database - 316 branches)

Outputs: herc_master_database.xlsx with 5 sheets
"""

import json
import csv
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = "/Users/vinuraabeysundara/Desktop/ICG/DOZR/SupplierBranches"

def extract_branch_number_from_oms(branch_name):
    """Extract branch number from OMS branch name like 'Herc Rentals (9422) - Kansas City MO'"""
    if not branch_name:
        return None
    match = re.search(r'\((\d+)\)', branch_name)
    if match:
        return match.group(1)
    return None

def extract_state_from_oms(branch_name):
    """Extract state from OMS branch name like 'Herc Rentals (9422) - Kansas City MO'"""
    if not branch_name:
        return None
    # Try to find 2-letter state code at end
    match = re.search(r'\s([A-Z]{2})$', branch_name.strip())
    if match:
        return match.group(1)
    return None

def extract_city_from_oms(branch_name):
    """Extract city from OMS branch name"""
    if not branch_name:
        return None
    # Format: "Herc Rentals (9422) - Kansas City MO"
    match = re.search(r'-\s*(.+?)\s+[A-Z]{2}$', branch_name)
    if match:
        return match.group(1).strip()
    return None

def load_web_scrape_xlsx():
    """Load branches from herc_complete_branch_database.xlsx - Web Branches sheet"""
    filepath = f"{BASE_DIR}/herc_complete_branch_database.xlsx"
    branches = []

    try:
        wb = load_workbook(filepath, read_only=True)
        ws = wb['Web Branches']  # Specifically load the Web Branches sheet

        # Column mapping based on actual structure:
        # 0: Branch #, 1: Branch Name, 2: Address, 3: City, 4: State, 5: Zip, 6: Phone, 7: Services, 8: Relevant to DOZR

        # Read data rows (skip header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue

            branch = {
                'branch_number': str(row[0] or '').strip(),
                'branch_name': str(row[1] or '').strip(),
                'address': str(row[2] or '').strip(),
                'city': str(row[3] or '').strip(),
                'state': str(row[4] or '').strip(),
                'zip': str(row[5] or '').strip(),
                'phone': str(row[6] or '').strip(),
                'services': str(row[7] or '').strip(),
                'source': 'web_scrape'
            }

            if branch['branch_number']:
                branches.append(branch)

        wb.close()
        print(f"Loaded {len(branches)} branches from web scrape xlsx")

    except Exception as e:
        print(f"Error loading xlsx: {e}")
        import traceback
        traceback.print_exc()

    return branches

def load_gap_fill_json():
    """Load branches from herc_gap_fill_branches.json"""
    filepath = f"{BASE_DIR}/herc_gap_fill_branches.json"
    branches = []

    try:
        with open(filepath, 'r') as f:
            data = json.load(f)

        for b in data.get('branches', []):
            branch = {
                'branch_number': str(b.get('branch_number', '')).strip(),
                'branch_name': str(b.get('branch_name', '')).strip(),
                'address': str(b.get('address', '')).strip(),
                'city': str(b.get('city', '')).strip(),
                'state': str(b.get('state', '')).strip(),
                'zip': str(b.get('zip', '')).strip(),
                'phone': str(b.get('phone', '')).strip(),
                'services': str(b.get('services', '')).strip(),
                'source': 'gap_fill'
            }

            if branch['branch_number']:
                branches.append(branch)

        print(f"Loaded {len(branches)} branches from gap fill JSON")

    except Exception as e:
        print(f"Error loading JSON: {e}")

    return branches

def load_oms_csv():
    """Load branches from OMS CSV"""
    filepath = f"{BASE_DIR}/Herc-Rentals-19-Feb-26.csv"
    branches = []

    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)

            for row in reader:
                branch_name = row.get('branchName', '')
                branch_number = extract_branch_number_from_oms(branch_name)

                if not branch_number:
                    continue

                # Parse address
                address_full = row.get('branchAddress', '')
                address_parts = [p.strip() for p in address_full.split(',')]

                street = address_parts[0] if len(address_parts) > 0 else ''
                city = address_parts[1] if len(address_parts) > 1 else extract_city_from_oms(branch_name)
                state = ''
                zip_code = ''

                # Look for state and zip in address parts
                for part in address_parts:
                    part = part.strip()
                    # Check if it's a state name or abbreviation
                    if len(part) == 2 and part.isalpha():
                        state = part
                    # Check for zip code
                    zip_match = re.search(r'(\d{5}(-\d{4})?)', part)
                    if zip_match:
                        zip_code = zip_match.group(1)

                # If state not found, try extracting from branch name
                if not state:
                    state = extract_state_from_oms(branch_name) or ''

                branch = {
                    'branch_number': branch_number,
                    'branch_name': branch_name,
                    'address': street,
                    'city': city or '',
                    'state': state,
                    'zip': zip_code,
                    'phone': row.get('branchPhoneNumber', ''),
                    'services': '',
                    'oms_enabled': row.get('omsEnabled', 'FALSE').upper() == 'TRUE',
                    'oms_email': row.get('omsEmailAddresses', ''),
                    'source': 'oms'
                }

                branches.append(branch)

        print(f"Loaded {len(branches)} branches from OMS CSV")

    except Exception as e:
        print(f"Error loading CSV: {e}")

    return branches

def combine_and_dedupe_web_data(web_scrape, gap_fill):
    """Combine web scrape and gap fill, dedupe by branch number"""
    combined = {}

    # Add web scrape first
    for b in web_scrape:
        num = b['branch_number']
        if num:
            combined[num] = b

    web_count = len(combined)

    # Add gap fill (overwrites if exists, since gap fill is newer)
    for b in gap_fill:
        num = b['branch_number']
        if num:
            if num in combined:
                # Mark as found in both
                combined[num]['source'] = 'both'
            else:
                combined[num] = b

    print(f"Combined web data: {web_count} from web scrape + {len(gap_fill)} from gap fill = {len(combined)} unique branches")

    return combined

def create_master_database(web_combined, oms_branches):
    """Create master database with all required sheets"""

    # Convert OMS to dict by branch number
    oms_dict = {}
    for b in oms_branches:
        num = b['branch_number']
        if num:
            oms_dict[num] = b

    # Find all unique branches
    all_numbers = set(web_combined.keys()) | set(oms_dict.keys())

    # Categorize
    all_unique = []
    new_to_add = []  # On website but NOT in OMS
    oms_only = []    # In OMS but NOT on website

    for num in sorted(all_numbers):
        in_web = num in web_combined
        in_oms = num in oms_dict

        # Create unified branch record
        if in_web:
            branch = web_combined[num].copy()
            branch['in_web'] = True
            branch['in_oms'] = in_oms
            if in_oms:
                branch['oms_enabled'] = oms_dict[num].get('oms_enabled', False)
                branch['oms_email'] = oms_dict[num].get('oms_email', '')
        else:
            branch = oms_dict[num].copy()
            branch['in_web'] = False
            branch['in_oms'] = True

        all_unique.append(branch)

        if in_web and not in_oms:
            new_to_add.append(branch)
        elif in_oms and not in_web:
            oms_only.append(branch)

    # Create workbook
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    new_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    oms_fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: All Unique Branches
    ws1 = wb.active
    ws1.title = "All Unique Branches"
    headers = ['Branch #', 'Branch Name', 'Address', 'City', 'State', 'ZIP', 'Phone', 'Services', 'In Web', 'In OMS', 'OMS Enabled', 'Source']
    ws1.append(headers)

    for col in range(1, len(headers) + 1):
        ws1.cell(row=1, column=col).font = header_font
        ws1.cell(row=1, column=col).fill = header_fill
        ws1.cell(row=1, column=col).border = thin_border

    for b in all_unique:
        row = [
            b.get('branch_number', ''),
            b.get('branch_name', ''),
            b.get('address', ''),
            b.get('city', ''),
            b.get('state', ''),
            b.get('zip', ''),
            b.get('phone', ''),
            b.get('services', ''),
            'Yes' if b.get('in_web') else 'No',
            'Yes' if b.get('in_oms') else 'No',
            'Yes' if b.get('oms_enabled') else 'No',
            b.get('source', '')
        ]
        ws1.append(row)

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 15
    ws1.column_dimensions['B'].width = 40
    ws1.column_dimensions['C'].width = 35

    # Sheet 2: New Branches to Add to OMS
    ws2 = wb.create_sheet("New to Add to OMS")
    headers2 = ['Branch #', 'Branch Name', 'Address', 'City', 'State', 'ZIP', 'Phone', 'Services', 'Source']
    ws2.append(headers2)

    for col in range(1, len(headers2) + 1):
        ws2.cell(row=1, column=col).font = header_font
        ws2.cell(row=1, column=col).fill = PatternFill(start_color="43A047", end_color="43A047", fill_type="solid")
        ws2.cell(row=1, column=col).border = thin_border

    for b in new_to_add:
        row = [
            b.get('branch_number', ''),
            b.get('branch_name', ''),
            b.get('address', ''),
            b.get('city', ''),
            b.get('state', ''),
            b.get('zip', ''),
            b.get('phone', ''),
            b.get('services', ''),
            b.get('source', '')
        ]
        ws2.append(row)

    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 15
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 35

    # Sheet 3: OMS Only (not on website)
    ws3 = wb.create_sheet("OMS Only")
    headers3 = ['Branch #', 'Branch Name', 'Address', 'City', 'State', 'ZIP', 'Phone', 'OMS Enabled', 'OMS Email']
    ws3.append(headers3)

    for col in range(1, len(headers3) + 1):
        ws3.cell(row=1, column=col).font = header_font
        ws3.cell(row=1, column=col).fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        ws3.cell(row=1, column=col).border = thin_border

    for b in oms_only:
        row = [
            b.get('branch_number', ''),
            b.get('branch_name', ''),
            b.get('address', ''),
            b.get('city', ''),
            b.get('state', ''),
            b.get('zip', ''),
            b.get('phone', ''),
            'Yes' if b.get('oms_enabled') else 'No',
            b.get('oms_email', '')
        ]
        ws3.append(row)

    for col in range(1, len(headers3) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 15
    ws3.column_dimensions['B'].width = 40
    ws3.column_dimensions['I'].width = 50

    # Sheet 4: State Summary
    ws4 = wb.create_sheet("State Summary")

    # Count by state
    state_counts = {}
    for b in all_unique:
        state = b.get('state', 'Unknown') or 'Unknown'
        if state not in state_counts:
            state_counts[state] = {'total': 0, 'web': 0, 'oms': 0, 'new': 0, 'oms_only': 0}
        state_counts[state]['total'] += 1
        if b.get('in_web'):
            state_counts[state]['web'] += 1
        if b.get('in_oms'):
            state_counts[state]['oms'] += 1
        if b.get('in_web') and not b.get('in_oms'):
            state_counts[state]['new'] += 1
        if b.get('in_oms') and not b.get('in_web'):
            state_counts[state]['oms_only'] += 1

    headers4 = ['State', 'Total Unique', 'On Website', 'In OMS', 'New to Add', 'OMS Only']
    ws4.append(headers4)

    for col in range(1, len(headers4) + 1):
        ws4.cell(row=1, column=col).font = header_font
        ws4.cell(row=1, column=col).fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
        ws4.cell(row=1, column=col).border = thin_border

    for state in sorted(state_counts.keys()):
        counts = state_counts[state]
        ws4.append([state, counts['total'], counts['web'], counts['oms'], counts['new'], counts['oms_only']])

    # Add totals row
    totals = ['TOTAL', len(all_unique),
              sum(1 for b in all_unique if b.get('in_web')),
              sum(1 for b in all_unique if b.get('in_oms')),
              len(new_to_add), len(oms_only)]
    ws4.append(totals)
    last_row = ws4.max_row
    for col in range(1, len(headers4) + 1):
        ws4.cell(row=last_row, column=col).font = Font(bold=True)

    for col in range(1, len(headers4) + 1):
        ws4.column_dimensions[get_column_letter(col)].width = 15

    # Sheet 5: Summary
    ws5 = wb.create_sheet("Summary")

    summary_data = [
        ['HERC RENTALS MASTER DATABASE SUMMARY', ''],
        ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M')],
        ['', ''],
        ['SOURCE DATA', ''],
        ['Web Scrape (herc_complete_branch_database.xlsx)', len([b for b in all_unique if b.get('source') in ['web_scrape', 'both']])],
        ['Gap Fill (herc_gap_fill_branches.json)', len([b for b in all_unique if b.get('source') in ['gap_fill', 'both']])],
        ['OMS CSV (Herc-Rentals-19-Feb-26.csv)', len(oms_branches)],
        ['', ''],
        ['DEDUPLICATION RESULTS', ''],
        ['Total Unique Branches (All Sources)', len(all_unique)],
        ['Branches on Website (Web + Gap Fill)', len(web_combined)],
        ['Branches in OMS', len(oms_dict)],
        ['', ''],
        ['CROSS-REFERENCE ANALYSIS', ''],
        ['NEW - On Website, NOT in OMS', len(new_to_add)],
        ['MATCHED - In Both Website & OMS', len(all_unique) - len(new_to_add) - len(oms_only)],
        ['OMS ONLY - In OMS, NOT on Website', len(oms_only)],
        ['', ''],
        ['OMS STATUS', ''],
        ['Currently OMS Enabled', sum(1 for b in oms_branches if b.get('oms_enabled'))],
        ['Not OMS Enabled', sum(1 for b in oms_branches if not b.get('oms_enabled'))],
    ]

    for row in summary_data:
        ws5.append(row)

    # Style the summary
    ws5.cell(row=1, column=1).font = Font(bold=True, size=14)
    for row_num in [4, 9, 14, 19]:
        ws5.cell(row=row_num, column=1).font = Font(bold=True)

    ws5.column_dimensions['A'].width = 45
    ws5.column_dimensions['B'].width = 15

    # Save workbook
    output_path = f"{BASE_DIR}/herc_master_database.xlsx"
    wb.save(output_path)

    return {
        'all_unique': len(all_unique),
        'new_to_add': len(new_to_add),
        'oms_only': len(oms_only),
        'matched': len(all_unique) - len(new_to_add) - len(oms_only),
        'web_total': len(web_combined),
        'oms_total': len(oms_dict),
        'output_path': output_path
    }

def main():
    print("=" * 70)
    print("CREATING HERC RENTALS MASTER DATABASE")
    print("=" * 70)
    print()

    # Load all data sources
    print("Loading data sources...")
    web_scrape = load_web_scrape_xlsx()
    gap_fill = load_gap_fill_json()
    oms_data = load_oms_csv()

    print()

    # Combine web data
    print("Combining and deduplicating web data...")
    web_combined = combine_and_dedupe_web_data(web_scrape, gap_fill)

    print()

    # Create master database
    print("Creating master database...")
    results = create_master_database(web_combined, oms_data)

    print()
    print("=" * 70)
    print("FINAL COUNTS")
    print("=" * 70)
    print(f"Total Unique Branches:        {results['all_unique']}")
    print(f"On Website (Combined):        {results['web_total']}")
    print(f"In OMS Database:              {results['oms_total']}")
    print()
    print(f"NEW to Add to OMS:            {results['new_to_add']}")
    print(f"Matched (Both):               {results['matched']}")
    print(f"OMS Only (Not on Website):    {results['oms_only']}")
    print()
    print(f"Output saved to: {results['output_path']}")
    print("=" * 70)

if __name__ == "__main__":
    main()
