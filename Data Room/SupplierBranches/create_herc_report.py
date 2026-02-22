#!/usr/bin/env python3
"""
Create Herc Rentals comprehensive Excel report with OMS cross-referencing
"""

import json
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    os.system("pip3 install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = "/Users/vinuraabeysundara/Desktop/ICG/DOZR/SupplierBranches"

# Styles
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NEW_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
OMS_ONLY_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def load_json(filename):
    filepath = os.path.join(BASE_DIR, filename)
    if os.path.exists(filepath):
        with open(filepath, 'r') as f:
            return json.load(f)
    return None

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER

def main():
    print("Creating Herc Rentals Complete Report...")

    # Load data
    oms_data = load_json('herc_oms_branches.json')
    web_data = load_json('herc_web_branches.json')

    oms_branches = oms_data.get('branches', []) if oms_data else []
    web_branches = web_data.get('branches', []) if web_data else []

    print(f"OMS branches: {len(oms_branches)}")
    print(f"Web branches: {len(web_branches)}")

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: All Web Branches
    ws1 = wb.create_sheet("Web Branches")
    headers1 = ["Branch #", "Branch Name", "Address", "City", "State", "Zip", "Phone", "Services", "Relevant to DOZR"]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header(ws1, 1, len(headers1))

    # Sort web branches by state then city
    web_sorted = sorted(web_branches, key=lambda x: (x.get('state', ''), x.get('city', '')))
    for row, b in enumerate(web_sorted, 2):
        ws1.cell(row=row, column=1, value=b.get('branch_number', ''))
        ws1.cell(row=row, column=2, value=b.get('branch_name', ''))
        ws1.cell(row=row, column=3, value=b.get('address', ''))
        ws1.cell(row=row, column=4, value=b.get('city', ''))
        ws1.cell(row=row, column=5, value=b.get('state', ''))
        ws1.cell(row=row, column=6, value=b.get('zip', ''))
        ws1.cell(row=row, column=7, value=b.get('phone', ''))
        ws1.cell(row=row, column=8, value=b.get('services', ''))
        # Mark relevant services
        services = b.get('services', '').lower()
        relevant = 'Yes' if any(s in services for s in ['general', 'aerial', 'crane', 'earthmov', 'excavat']) else 'Check'
        ws1.cell(row=row, column=9, value=relevant)
        for col in range(1, len(headers1) + 1):
            ws1.cell(row=row, column=col).border = BORDER

    # Set column widths
    widths1 = [10, 40, 35, 20, 8, 10, 15, 30, 12]
    for col, w in enumerate(widths1, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws1.freeze_panes = 'A2'

    # Sheet 2: Cross-Reference with OMS
    ws2 = wb.create_sheet("Cross-Reference")
    headers2 = ["Branch #", "Branch Name", "City", "State", "Phone", "In OMS?", "On Website?", "Status"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, 1, len(headers2))

    # Build lookup sets
    oms_numbers = {b.get('branch_id', '').split('(')[-1].replace(')', '').strip()
                   for b in oms_branches if '(' in b.get('branch_id', '')}
    # Also extract branch numbers from names like "Herc Rentals (9321) - Charlotte NC"
    for b in oms_branches:
        name = b.get('branch_name', '')
        if '(' in name and ')' in name:
            num = name.split('(')[1].split(')')[0]
            if num.isdigit():
                oms_numbers.add(num)

    web_numbers = {b.get('branch_number', '') for b in web_branches}

    all_numbers = oms_numbers | web_numbers

    # Create combined view
    row = 2
    matched = 0
    new_to_add = 0
    oms_only = 0

    # Process OMS branches
    oms_by_number = {}
    for b in oms_branches:
        name = b.get('branch_name', '')
        if '(' in name and ')' in name:
            num = name.split('(')[1].split(')')[0]
            if num.isdigit():
                oms_by_number[num] = b

    web_by_number = {b.get('branch_number', ''): b for b in web_branches}

    for num in sorted(all_numbers):
        if not num:
            continue
        in_oms = num in oms_numbers or num in oms_by_number
        in_web = num in web_numbers

        # Get branch details
        if num in web_by_number:
            b = web_by_number[num]
            name = b.get('branch_name', '')
            city = b.get('city', '')
            state = b.get('state', '')
            phone = b.get('phone', '')
        elif num in oms_by_number:
            b = oms_by_number[num]
            name = b.get('branch_name', '')
            city = b.get('city', '')
            state = b.get('state', '')
            phone = b.get('phone', '')
        else:
            name = city = state = phone = ''

        ws2.cell(row=row, column=1, value=num)
        ws2.cell(row=row, column=2, value=name)
        ws2.cell(row=row, column=3, value=city)
        ws2.cell(row=row, column=4, value=state)
        ws2.cell(row=row, column=5, value=phone)
        ws2.cell(row=row, column=6, value="Yes" if in_oms else "No")
        ws2.cell(row=row, column=7, value="Yes" if in_web else "No")

        if in_oms and in_web:
            status = "MATCHED"
            ws2.cell(row=row, column=8).fill = MATCH_FILL
            matched += 1
        elif in_web and not in_oms:
            status = "NEW - ADD TO OMS"
            ws2.cell(row=row, column=8).fill = NEW_FILL
            new_to_add += 1
        else:
            status = "OMS ONLY"
            ws2.cell(row=row, column=8).fill = OMS_ONLY_FILL
            oms_only += 1

        ws2.cell(row=row, column=8, value=status)
        for col in range(1, len(headers2) + 1):
            ws2.cell(row=row, column=col).border = BORDER
        row += 1

    widths2 = [10, 45, 20, 8, 15, 10, 12, 18]
    for col, w in enumerate(widths2, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws2.freeze_panes = 'A2'

    # Sheet 3: New Branches to Add
    ws3 = wb.create_sheet("New to Add to OMS")
    headers3 = ["Branch #", "Branch Name", "Address", "City", "State", "Zip", "Phone", "Services"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    style_header(ws3, 1, len(headers3))

    row = 2
    for b in web_sorted:
        num = b.get('branch_number', '')
        if num and num not in oms_numbers and num not in oms_by_number:
            ws3.cell(row=row, column=1, value=num)
            ws3.cell(row=row, column=2, value=b.get('branch_name', ''))
            ws3.cell(row=row, column=3, value=b.get('address', ''))
            ws3.cell(row=row, column=4, value=b.get('city', ''))
            ws3.cell(row=row, column=5, value=b.get('state', ''))
            ws3.cell(row=row, column=6, value=b.get('zip', ''))
            ws3.cell(row=row, column=7, value=b.get('phone', ''))
            ws3.cell(row=row, column=8, value=b.get('services', ''))
            for col in range(1, len(headers3) + 1):
                ws3.cell(row=row, column=col).border = BORDER
            row += 1

    widths3 = [10, 40, 35, 20, 8, 10, 15, 30]
    for col, w in enumerate(widths3, 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws3.freeze_panes = 'A2'

    # Sheet 4: State Summary
    ws4 = wb.create_sheet("State Summary")
    headers4 = ["State", "Website Count", "OMS Count", "Matched", "New to Add"]
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=col, value=h)
    style_header(ws4, 1, len(headers4))

    # Count by state
    web_by_state = {}
    for b in web_branches:
        state = b.get('state', 'Unknown')
        web_by_state[state] = web_by_state.get(state, 0) + 1

    oms_by_state = {}
    for b in oms_branches:
        state = b.get('state', 'Unknown')
        oms_by_state[state] = oms_by_state.get(state, 0) + 1

    all_states = sorted(set(web_by_state.keys()) | set(oms_by_state.keys()))
    row = 2
    for state in all_states:
        if not state:
            continue
        web_count = web_by_state.get(state, 0)
        oms_count = oms_by_state.get(state, 0)
        ws4.cell(row=row, column=1, value=state)
        ws4.cell(row=row, column=2, value=web_count)
        ws4.cell(row=row, column=3, value=oms_count)
        ws4.cell(row=row, column=4, value=min(web_count, oms_count))
        ws4.cell(row=row, column=5, value=max(0, web_count - oms_count))
        for col in range(1, len(headers4) + 1):
            ws4.cell(row=row, column=col).border = BORDER
        row += 1

    # Totals
    ws4.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws4.cell(row=row, column=2, value=len(web_branches)).font = Font(bold=True)
    ws4.cell(row=row, column=3, value=len(oms_branches)).font = Font(bold=True)
    ws4.cell(row=row, column=4, value=matched).font = Font(bold=True)
    ws4.cell(row=row, column=5, value=new_to_add).font = Font(bold=True)

    widths4 = [10, 15, 12, 10, 12]
    for col, w in enumerate(widths4, 1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # Sheet 5: Summary
    ws5 = wb.create_sheet("Summary", 0)
    ws5.cell(row=1, column=1, value="Herc Rentals Branch Database Report").font = Font(bold=True, size=14)
    ws5.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws5.cell(row=4, column=1, value="Data Sources:").font = Font(bold=True)
    ws5.cell(row=5, column=1, value=f"  OMS Export: {len(oms_branches)} branches")
    ws5.cell(row=6, column=1, value=f"  Website Extraction: {len(web_branches)} branches (partial - 4 regions)")
    ws5.cell(row=8, column=1, value="Cross-Reference Results:").font = Font(bold=True)
    ws5.cell(row=9, column=1, value=f"  Matched (in both): {matched}")
    ws5.cell(row=10, column=1, value=f"  New to Add to OMS: {new_to_add}")
    ws5.cell(row=11, column=1, value=f"  OMS Only (verify): {oms_only}")
    ws5.cell(row=13, column=1, value="Regions Searched on Website:").font = Font(bold=True)
    ws5.cell(row=14, column=1, value="  Los Angeles CA, Houston TX, New York NY/NJ, Atlanta GA")
    ws5.cell(row=16, column=1, value="Notes:").font = Font(bold=True)
    ws5.cell(row=17, column=1, value="  - Website has ~800+ branches total (only partial extraction)")
    ws5.cell(row=18, column=1, value="  - OMS data is authoritative for existing DOZR integration")
    ws5.cell(row=19, column=1, value="  - Only 3 of 316 OMS branches are currently enabled")

    ws5.column_dimensions['A'].width = 60

    # Save
    output_file = os.path.join(BASE_DIR, "herc_complete_branch_database.xlsx")
    wb.save(output_file)

    print(f"\nReport saved to: {output_file}")
    print(f"\nSummary:")
    print(f"  OMS branches: {len(oms_branches)}")
    print(f"  Web branches: {len(web_branches)}")
    print(f"  Matched: {matched}")
    print(f"  New to add: {new_to_add}")
    print(f"  OMS only: {oms_only}")

if __name__ == "__main__":
    main()
