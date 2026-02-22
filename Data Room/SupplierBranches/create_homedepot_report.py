#!/usr/bin/env python3
"""
Create Home Depot Branch Database Excel Report
From extracted Priority 1 state data
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Read extracted data
with open('/Users/vinuraabeysundara/Desktop/ICG/DOZR/SupplierBranches/homedepot_extraction_progress.json', 'r') as f:
    data = json.load(f)

# Read OMS data for cross-reference
import csv
oms_branches = []
try:
    with open('/Users/vinuraabeysundara/Desktop/ICG/DOZR/SupplierBranches/The-Home-Depot-Rental-19-Feb-26.csv', 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            oms_branches.append({
                'branch_name': row.get('branchName', ''),
                'address': row.get('branchAddress', ''),
                'phone': row.get('branchPhoneNumber', ''),
                'state': row.get('branchAddress', '').split(',')[-1].strip().split()[0] if row.get('branchAddress') else '',
                'oms_enabled': row.get('omsEnabled', '') == 'TRUE'
            })
except FileNotFoundError:
    print("OMS file not found, proceeding without cross-reference")

# Create workbook
wb = openpyxl.Workbook()

# Styles
header_fill = PatternFill(start_color="F96302", end_color="F96302", fill_type="solid")  # Home Depot orange
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ===== Sheet 1: All Branches by State =====
ws1 = wb.active
ws1.title = "All Branches by State"

headers = ["State", "Store Name", "Address", "City", "Zip", "Phone", "Has Rentals"]
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

row_idx = 2
for state_code, state_data in sorted(data['states_data'].items()):
    for store in state_data['stores']:
        ws1.cell(row=row_idx, column=1, value=state_code).border = thin_border
        ws1.cell(row=row_idx, column=2, value=store['store_name']).border = thin_border
        ws1.cell(row=row_idx, column=3, value=store['address']).border = thin_border
        ws1.cell(row=row_idx, column=4, value=store['city']).border = thin_border
        ws1.cell(row=row_idx, column=5, value=store['zip']).border = thin_border
        ws1.cell(row=row_idx, column=6, value=store['phone']).border = thin_border
        ws1.cell(row=row_idx, column=7, value="Yes" if store.get('has_rentals', True) else "No").border = thin_border
        row_idx += 1

ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 35
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 10
ws1.column_dimensions['F'].width = 18
ws1.column_dimensions['G'].width = 12

# ===== Sheet 2: State Summary =====
ws2 = wb.create_sheet("State Summary")

headers2 = ["State Code", "State Name", "Total Stores", "With Rentals", "OMS Coverage Status"]
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Get OMS states
oms_states = set(b['state'] for b in oms_branches if b['state'])

row_idx = 2
for state_code, state_data in sorted(data['states_data'].items()):
    ws2.cell(row=row_idx, column=1, value=state_code).border = thin_border
    ws2.cell(row=row_idx, column=2, value=state_data['state_name']).border = thin_border
    ws2.cell(row=row_idx, column=3, value=state_data['total_stores']).border = thin_border
    ws2.cell(row=row_idx, column=4, value=state_data['total_stores']).border = thin_border  # All have rentals

    oms_status = "In OMS" if state_code in oms_states else "NOT in OMS (Priority 1)"
    cell = ws2.cell(row=row_idx, column=5, value=oms_status)
    cell.border = thin_border
    if "NOT" in oms_status:
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    row_idx += 1

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 25

# ===== Sheet 3: Summary Report =====
ws3 = wb.create_sheet("Summary Report")

summary_data = [
    ["Home Depot Equipment Rental - Priority 1 States Analysis", ""],
    ["Report Date", datetime.now().strftime("%Y-%m-%d")],
    ["", ""],
    ["EXTRACTION SUMMARY", ""],
    ["Total States Extracted", data['summary']['states_completed']],
    ["Total Stores Found", data['summary']['total_stores']],
    ["States Remaining", data['summary']['states_remaining']],
    ["", ""],
    ["PRIORITY 1 STATES (Zero OMS Coverage)", ""],
    ["Total Priority 1 States", 18],
    ["States Completed", data['summary']['states_completed']],
    ["", ""],
    ["EXTRACTED STATES BREAKDOWN", ""],
]

# Add state breakdown
for state_code, state_data in sorted(data['states_data'].items()):
    summary_data.append([f"  {state_data['state_name']} ({state_code})", f"{state_data['total_stores']} stores"])

summary_data.extend([
    ["", ""],
    ["REMAINING PRIORITY 1 STATES", ""],
    ["  DC, IA, ID, KY, ME, MN, MT, ND, NE, NV, SD, WV", "12 states remaining"],
    ["", ""],
    ["NOTES", ""],
    ["", "All Home Depot stores in extracted states have Rental services"],
    ["", "Large Equipment includes: Mini Excavators, Skid Steers, Scissor Lifts, Boom Lifts"],
    ["", "These Priority 1 states have ZERO coverage in current OMS"],
    ["", "Recommend onboarding these locations for equipment rental coverage"],
])

for row_idx, (label, value) in enumerate(summary_data, 1):
    cell1 = ws3.cell(row=row_idx, column=1, value=label)
    cell2 = ws3.cell(row=row_idx, column=2, value=value)
    if row_idx == 1:
        cell1.font = Font(bold=True, size=14)
    elif label and not label.startswith("  ") and label != "":
        cell1.font = Font(bold=True)

ws3.column_dimensions['A'].width = 50
ws3.column_dimensions['B'].width = 30

# ===== Sheet 4: Cross-Reference with OMS =====
ws4 = wb.create_sheet("OMS Cross-Reference")

headers4 = ["State", "Store Name", "Address", "In Extracted Data", "In OMS", "Gap"]
for col, header in enumerate(headers4, 1):
    cell = ws4.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

row_idx = 2
for state_code, state_data in sorted(data['states_data'].items()):
    for store in state_data['stores']:
        ws4.cell(row=row_idx, column=1, value=state_code).border = thin_border
        ws4.cell(row=row_idx, column=2, value=store['store_name']).border = thin_border
        ws4.cell(row=row_idx, column=3, value=f"{store['address']}, {store['city']}, {store['state']} {store['zip']}").border = thin_border
        ws4.cell(row=row_idx, column=4, value="Yes").border = thin_border
        ws4.cell(row=row_idx, column=5, value="No").border = thin_border  # Priority 1 = not in OMS
        cell = ws4.cell(row=row_idx, column=6, value="NEW - Not in OMS")
        cell.border = thin_border
        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        row_idx += 1

ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 25
ws4.column_dimensions['C'].width = 50
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 10
ws4.column_dimensions['F'].width = 18

# Save workbook
output_path = '/Users/vinuraabeysundara/Desktop/ICG/DOZR/SupplierBranches/homedepot_priority1_branch_database.xlsx'
wb.save(output_path)

print(f"Created: {output_path}")
print(f"States extracted: {data['summary']['states_completed']}")
print(f"Total stores: {data['summary']['total_stores']}")
print("\nStates included:")
for state_code, state_data in sorted(data['states_data'].items()):
    print(f"  {state_data['state_name']}: {state_data['total_stores']} stores")
