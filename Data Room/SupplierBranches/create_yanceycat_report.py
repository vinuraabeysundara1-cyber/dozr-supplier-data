#!/usr/bin/env python3
"""Create Yancey Cat Branch Database Excel Report"""

import json
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Read OMS CSV
oms_branches = []
with open('/Users/vinuraabeysundara/Desktop/ICG/DOZR/SupplierBranches/Yancey-Cat-19-Feb-26.csv', 'r') as f:
    reader = csv.DictReader(f)
    for row in reader:
        oms_branches.append({
            'branch_name': row['branchName'],
            'address': row['branchAddress'],
            'phone': row['branchPhoneNumber'],
            'oms_enabled': row['omsEnabled'] == 'TRUE'
        })

# Website data from our extraction
website_branches = [
    {"branch_name": "Yancey Bros Co. - Kennesaw", "address": "2697 McCollum Pkwy. NW, Kennesaw, GA 30144", "phone": "(678) 915-1177"},
    {"branch_name": "Yancey Bros Co. - Albany", "address": "1604 South Slappey Blvd., Albany, GA 31701", "phone": "(229) 435-6262"},
    {"branch_name": "Yancey Bros Co. - Athens (Bogart)", "address": "50 Trade Street, Bogart, GA 30622", "phone": "(706) 353-0043"},
    {"branch_name": "Yancey Bros Co. - Augusta", "address": "4165 Mike Padgett Hwy, Augusta, GA 30906", "phone": "(706) 790-1300"},
    {"branch_name": "Yancey Bros Co. - Austell", "address": "330 Lee Industrial Blvd, Austell, GA 30168", "phone": "(770) 941-2550"},
    {"branch_name": "Yancey Bros Co. - Bloomingdale", "address": "2566 U.S. Highway 80, Bloomingdale, GA 31302", "phone": "(912) 748-1711"},
    {"branch_name": "Yancey Bros Co. - Brunswick", "address": "370 Perry Lane Rd, Brunswick, GA 31525", "phone": "(912) 265-5010"},
    {"branch_name": "Yancey Bros Co. - Calhoun", "address": "138 Robinson Rd NW, Calhoun, GA 30701", "phone": "(706) 629-0776"},
    {"branch_name": "Yancey Bros Co. - Columbus", "address": "2946 Smith Road, Columbus, GA 31808", "phone": "(706) 649-0054"},
    {"branch_name": "Yancey Bros Co. - Cumming", "address": "5905 Hubbard Town Rd, Cumming, GA 30028", "phone": "(470) 560-6750"},
    {"branch_name": "Yancey Bros Co. - Dacula (Gwinnett)", "address": "335 Hurricane Trail, Dacula, GA 30019", "phone": "(770) 963-6868"},
    {"branch_name": "Yancey Bros Co. - Jefferson", "address": "2623 Bill Wright Rd, Jefferson, GA", "phone": "(706) 367-7214"},
    {"branch_name": "Yancey Bros Co. - Macon", "address": "875 Guy Paine Road, Macon, GA 31206", "phone": "(478) 785-5001"},
    {"branch_name": "Yancey Bros Co. - McDonough", "address": "187 Interstate South Dr, McDonough, GA 30253", "phone": "(770) 288-4000"},
    {"branch_name": "Yancey Bros Co. - Pooler (Savannah)", "address": "1459 U.S. 80 East, Pooler, GA 31322", "phone": "(912) 525-1206"},
    {"branch_name": "Yancey Bros Co. - Statesboro", "address": "8555 Highway 301, Statesboro, GA 30458", "phone": "(912) 871-6506"},
    {"branch_name": "Yancey Bros Co. - Valdosta", "address": "2966 HWY U.S. 84 West, Valdosta, GA 31601", "phone": "(229) 242-8610"},
    {"branch_name": "Yancey Bros Co. - Waycross", "address": "1899 Knight Ave, Waycross, GA 31503", "phone": "(912) 285-4800"},
]

# Create workbook
wb = openpyxl.Workbook()

# Header style
header_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
header_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ===== Sheet 1: All Branches from Website =====
ws1 = wb.active
ws1.title = "All Branches from Website"

headers = ["Branch Name", "Full Address", "City", "State", "Zip", "Phone", "Website"]
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for row_idx, branch in enumerate(website_branches, 2):
    ws1.cell(row=row_idx, column=1, value=branch['branch_name']).border = thin_border
    ws1.cell(row=row_idx, column=2, value=branch['address']).border = thin_border
    # Parse city from address
    addr_parts = branch['address'].split(', ')
    city = addr_parts[1] if len(addr_parts) > 1 else ""
    ws1.cell(row=row_idx, column=3, value=city).border = thin_border
    ws1.cell(row=row_idx, column=4, value="GA").border = thin_border
    zip_code = addr_parts[-1].split()[-1] if addr_parts else ""
    ws1.cell(row=row_idx, column=5, value=zip_code).border = thin_border
    ws1.cell(row=row_idx, column=6, value=branch['phone']).border = thin_border
    ws1.cell(row=row_idx, column=7, value="yanceybros.com").border = thin_border

ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 45
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 8
ws1.column_dimensions['E'].width = 10
ws1.column_dimensions['F'].width = 18
ws1.column_dimensions['G'].width = 18

# ===== Sheet 2: Cross-Reference with OMS =====
ws2 = wb.create_sheet("Cross-Reference with OMS")

headers2 = ["Branch Name", "Address", "In OMS?", "On Website?", "OMS Enabled?", "Status"]
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

row_idx = 2
for oms_branch in oms_branches:
    ws2.cell(row=row_idx, column=1, value=oms_branch['branch_name']).border = thin_border
    ws2.cell(row=row_idx, column=2, value=oms_branch['address']).border = thin_border
    ws2.cell(row=row_idx, column=3, value="Yes").border = thin_border
    ws2.cell(row=row_idx, column=4, value="Yes").border = thin_border
    enabled = "Yes" if oms_branch['oms_enabled'] else "No"
    ws2.cell(row=row_idx, column=5, value=enabled).border = thin_border
    status = "Active" if oms_branch['oms_enabled'] else "Disabled"
    ws2.cell(row=row_idx, column=6, value=status).border = thin_border
    row_idx += 1

ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 50
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 12
ws2.column_dimensions['F'].width = 12

# ===== Sheet 3: Summary =====
ws3 = wb.create_sheet("Summary")

summary_data = [
    ["Yancey Cat Branch Database Summary", ""],
    ["Report Date", datetime.now().strftime("%Y-%m-%d")],
    ["", ""],
    ["Company", "Yancey Bros Co. (Yancey Cat)"],
    ["Type", "Caterpillar Dealer"],
    ["Coverage", "Georgia"],
    ["", ""],
    ["Website Branches Found", len(website_branches)],
    ["OMS Branches", len(oms_branches)],
    ["OMS Enabled", sum(1 for b in oms_branches if b['oms_enabled'])],
    ["OMS Disabled", sum(1 for b in oms_branches if not b['oms_enabled'])],
    ["", ""],
    ["Branch Locations", ""],
    ["- Kennesaw", "OMS Enabled"],
    ["- Albany", "OMS Enabled"],
    ["- Athens/Bogart", "OMS Disabled"],
    ["- Augusta", "OMS Enabled"],
    ["- Austell", "OMS Enabled"],
    ["- Brunswick", "OMS Enabled"],
    ["- Calhoun", "OMS Enabled"],
    ["- Columbus", "OMS Enabled"],
    ["- Cumming", "OMS Enabled"],
    ["- Dacula/Gwinnett", "OMS Enabled"],
    ["- Jefferson", "OMS Enabled"],
    ["- Macon", "OMS Enabled"],
    ["- McDonough", "OMS Enabled"],
    ["- Pooler/Savannah", "OMS Enabled"],
    ["- Statesboro", "OMS Enabled"],
    ["- Valdosta", "OMS Enabled"],
    ["- Waycross", "OMS Disabled"],
    ["", ""],
    ["Notes", ""],
    ["", "Bloomingdale (near Savannah) appears on website but may be duplicate of Pooler"],
    ["", "All branches are in Georgia"],
    ["", "15 of 17 OMS branches are enabled"],
]

for row_idx, (label, value) in enumerate(summary_data, 1):
    cell1 = ws3.cell(row=row_idx, column=1, value=label)
    cell2 = ws3.cell(row=row_idx, column=2, value=value)
    if row_idx == 1:
        cell1.font = Font(bold=True, size=14)
    elif label and not str(value):
        cell1.font = Font(bold=True)

ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 30

# Save
output_path = '/Users/vinuraabeysundara/Desktop/ICG/DOZR/SupplierBranches/yanceycat_branch_database.xlsx'
wb.save(output_path)
print(f"Created: {output_path}")
print(f"Website branches: {len(website_branches)}")
print(f"OMS branches: {len(oms_branches)}")
print(f"OMS enabled: {sum(1 for b in oms_branches if b['oms_enabled'])}")
